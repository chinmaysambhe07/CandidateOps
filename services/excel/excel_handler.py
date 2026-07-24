"""
Excel handling module for CandidateOps.
Manages reading candidate data from Excel templates and writing results back.
"""


from __future__ import annotations
from pathlib import Path
from typing import List, Optional, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models.candidate import Candidate
from utils.exceptions import ExcelError
from utils.logging_setup import get_logger
from infrastructure.config.settings import settings


logger = get_logger(__name__)


class ExcelHandler:
    """
    Handles Excel file operations for candidate tracking.
    """

    def __init__(self, template_path: Optional[Path] = None,
                 output_path: Optional[Path] = None):
        """
        Initialize Excel handler.

        Args:
            template_path: Path to Excel template file.
            output_path: Path for output Excel file.
        """
        self.template_path = template_path or settings.excel.template_path
        self.output_path = output_path or settings.excel.output_path
        self.sheet_name = settings.excel.sheet_name
        self.id_column = settings.excel.id_column

        # Ensure output directory exists
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        # Initialize workbook if template exists, otherwise create new
        if self.template_path.exists():
            self.workbook = load_workbook(self.template_path)
            logger.info(f"Loaded Excel template from {self.template_path}")
        else:
            self.workbook = self._create_new_workbook()
            logger.info(f"Created new Excel workbook (template not found at {self.template_path})")

        # Ensure the sheet exists
        if self.sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(self.sheet_name)
            logger.info(f"Created sheet '{self.sheet_name}'")

        self.worksheet = self.workbook[self.sheet_name]

    def _create_new_workbook(self) -> pd.ExcelFile:
        """
        Create a new Excel workbook with default structure.

        Returns:
            New workbook instance.
        """
        # Create a simple DataFrame with column headers
        df = pd.DataFrame(columns=[
            "CandidateID", "Name", "Email", "Phone", "City", "CurrentPosition",
            "YearsExperience", "Education", "Skills", "ApplicationDate", "Status",
            "ResumePath", "CoverLetterPath", "AttachmentsCount", "Notes"
        ])

        # Save to create the file
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=self.sheet_name, index=False)

        return load_workbook(self.output_path)

    def read_candidates(self) -> List[Candidate]:
        """
        Read candidate data from Excel file.

        Returns:
            List of Candidate objects.

        Raises:
            ExcelError: If there are issues reading the Excel file.
        """
        try:
            if not self.output_path.exists():
                logger.warning(f"Output Excel file not found at {self.output_path}. Returning empty list.")
                return []

            # Read data from Excel
            df = pd.read_excel(
                self.output_path,
                sheet_name=self.sheet_name,
                dtype={'CandidateID': str}  # Ensure ID is treated as string
            )

            # Convert DataFrame rows to Candidate objects
            candidates = []
            for _, row in df.iterrows():
                try:
                    # Skip empty rows
                    if pd.isna(row.get('CandidateID')) or str(row.get('CandidateID')).strip() == '':
                        continue

                    candidate = Candidate.from_dict(row.to_dict())
                    candidates.append(candidate)
                    logger.debug(f"Loaded {len(candidates)} candidates from Excel")
                return candidates

        except Exception as e:
            logger.error(f"Failed to read candidates from Excel: {str(e)}")
            raise ExcelError(f"Failed to read candidates from Excel: {str(e)}")

    def write_candidates(self, candidates: List[Candidate]) -> None:
        """
        Write candidate data to Excel file.

        Args:
            candidates: List of Candidate objects to write.

        Raises:
            ExcelError: If there are issues writing to the Excel file.
        """
        try:
            # Convert candidates to DataFrame
            data = [candidate.to_dict() for candidate in candidates]
            df = pd.DataFrame(data)

            # Ensure output directory exists
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

            # Write to Excel
            with pd.ExcelWriter(
                self.output_path,
                engine='openpyxl',
                mode='w' if not self.output_path.exists() else 'a',
                if_sheet_exists='replace'
            ) as writer:
                df.to_excel(writer, sheet_name=self.sheet_name, index=False)

            # Apply formatting
            self._apply_excel_formatting()

            logger.info(f"Successfully wrote {len(candidates)} candidates to Excel: {self.output_path}")

        except Exception as e:
            logger.error(f"Failed to write candidates to Excel: {str(e)}")
            raise ExcelError(f"Failed to write candidates to Excel: {str(e)}")

    def _apply_excel_formatting(self) -> None:
        """Apply formatting to the Excel worksheet for better readability."""
        try:
            # Load workbook for formatting
            wb = load_workbook(self.output_path)
            ws = wb[self.sheet_name]

            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Apply header formatting
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save formatted workbook
            wb.save(self.output_path)
            logger.debug("Applied Excel formatting")

        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting: {str(e)}")
            # Don't raise exception as formatting failure shouldn't break the main functionality

    def get_next_available_row(self) -> int:
        """
        Get the next available row number for writing data.

        Returns:
            Row number (1-indexed) where new data should be written.
        """
        if self.worksheet.max_row == 1:
            # Only header exists
            return 2
        return self.worksheet.max_row + 1

    def find_candidate_by_id(self, candidate_id: str) -> Optional[Candidate]:
        """
        Find a candidate by their ID in the Excel file.

        Args:
            candidate_id: Candidate ID to search for.

        Returns:
            Candidate object if found, None otherwise.
        """
        try:
            candidates = self.read_candidates()
            for candidate in candidates:
                if candidate.candidate_id == candidate_id:
                    return candidate
            return None
        except Exception as e:
            logger.error(f"Error searching for candidate {candidate_id}: {str(e)}")
            return None

    def update_candidate(self, candidate: Candidate) -> bool:
        """
        Update an existing candidate in the Excel file.

        Args:
            candidate: Candidate object with updated data.

        Returns:
            True if candidate was updated, False if not found.
        """
        try:
            candidates = self.read_candidates()
            updated = False

            for i, existing_candidate in enumerate(candidates):
                if existing_candidate.candidate_id == candidate.candidate_id:
                    candidates[i] = candidate
                    updated = True
                    break

            if updated:
                self.write_candidates(candidates)
                logger.info(f"Updated candidate {candidate.candidate_id} in Excel")
            else:
                logger.warning(f"Candidate {candidate.candidate_id} not found for update")

            return updated

        except Exception as e:
            logger.error(f"Failed to update candidate {candidate.candidate_id}: {str(e)}")
            raise ExcelError(f"Failed to update candidate: {str(e)}")

    def add_candidate(self, candidate: Candidate) -> None:
        """
        Add a new candidate to the Excel file.

        Args:
            candidate: Candidate object to add.

        Raises:
            ExcelError: If candidate with same ID already exists.
        """
        try:
            # Check if candidate already exists
            existing = self.find_candidate_by_id(candidate.candidate_id)
            if existing:
                raise ExcelError(f"Candidate with ID {candidate.candidate_id} already exists")

            candidates = self.read_candidates()
            candidates.append(candidate)
            self.write_candidates(candidates)
            logger.info(f"Added new candidate {candidate.candidate_id} to Excel")

        except Exception as e:
            logger.error(f"Failed to add candidate {candidate.candidate_id}: {str(e)}")
            raise ExcelError(f"Failed to add candidate: {str(e)}")